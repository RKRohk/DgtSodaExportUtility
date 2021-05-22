import json
import openpyxl
import os
from pathlib import Path
import gzip
import shutil

mypath = Path('/home/rohit/Dropbox/Apps/DGT-SODA')

files = sorted(os.listdir(os.getcwd()),key = os.path.getmtime)
gzfiles = filter(lambda x: True if x.endswith('json.gz') else False, files)
gzfiles = list(gzfiles)
# print(gzfiles)
newestfile = gzfiles[-1]

with gzip.open(newestfile) as infile:  
    data = json.load(infile)
    wbTasks = openpyxl.Workbook()
    sheet = wbTasks.active
    sheet.title = "Imported Tasks"
    rowNum=2    
    colNum=0
        
    # for things in data["task"]:
    #     sheet.cell(row = rowNum,column = 6 if things['parentId'] == 0 else 7).value = (things['title'])
    #     sheet.cell(row = rowNum,column = 2).value = things['id']
    #     parentId = things['parentId']
    #     for things in data["task"]:
    #         if things["parentId"] == parentId and parentId != 0:
    #             sheet.cell(row = rowNum,column = 5).value = (things['title'])
    #     rowNum = rowNum + 1
    # wbTasks.save('NewTasks.xlsx')
    tasks = []
    tasks = data["task"]
    tasks = list(filter(lambda x: True if x['completed'] == 0 else False,tasks))
    for task in tasks:
        
        if task['completed'] == 0:
            sheet.cell(row = rowNum,column = 1).value = (task["completed"])
            sheet.cell(row = rowNum,column = 2).value = (task["dueDate"])
            sheet.cell(row = rowNum,column = 3).value = (task["dueDateMod"])
            sheet.cell(row = rowNum,column = 4).value = (task["dueTimeSet"])
            sheet.cell(row = rowNum,column = 5).value = (task["duration"])
            sheet.cell(row = rowNum,column = 6).value = (task["energy"])
            sheet.cell(row = rowNum,column = 7).value = (task["hasActiveLocations"])
            sheet.cell(row = rowNum,column = 8).value = (task["hasActiveReminders"])
            sheet.cell(row = rowNum,column = 9).value = (task["hideUntil"])
            sheet.cell(row = rowNum,column = 10).value = (task["importance"])
            sheet.cell(row = rowNum,column = 11).value = (task["preventAutoPurge"])
            sheet.cell(row = rowNum,column = 12).value = (task["priority"])
            sheet.cell(row = rowNum,column = 13).value = (task["repeatFrom"])
            sheet.cell(row = rowNum,column = 14).value = (task["retentionPolicy"])
            sheet.cell(row = rowNum,column = 15).value = (task["starred"])
            sheet.cell(row = rowNum,column = 16).value = (task["startDate"])
            sheet.cell(row = rowNum,column = 17).value = (task["startDateMod"])
            sheet.cell(row = rowNum,column = 18).value = (task["startTimeSet"])
            sheet.cell(row = rowNum,column = 19).value = (task["status"])
            sheet.cell(row = rowNum,column = 20).value = (str(task.get("timeZone","Asia/Kolkata")))
            sheet.cell(row = rowNum,column = 21).value = (task["type"])
            sheet.cell(row = rowNum,column = 22).value = (task["urgency"])
            sheet.cell(row = rowNum,column = 23).value = (task["accountId"])
            sheet.cell(row = rowNum,column = 24).value = (task["created"])
            sheet.cell(row = rowNum,column = 25).value = (task["deleted"])
            sheet.cell(row = rowNum,column = 26).value = (task["id"])
            sheet.cell(row = rowNum,column = 27).value = (task["indentation"])
            sheet.cell(row = rowNum,column = 28).value = (task["modified"])
            sheet.cell(row = rowNum,column = 29).value = (task["parentId"])
            sheet.cell(row = rowNum,column = 30).value = (task["position"])
            sheet.cell(row = rowNum,column = 31).value = (task["title"])
            sheet.cell(row = rowNum,column = 32).value = (task["uuid"])
            rowNum = rowNum + 1
    
    r = 1
    c = 1
    for items in tasks[0]:
        sheet.cell(row=r,column=c).value = items
        print(tasks[0][items])
        c += 1
        
    wbTasks.save('TasksListImport.xlsx')
    print('saved!')



#putting it all in excel
# wb= openpyxl.Workbook()
# actSheet=wb.active
# sheet.title = 'Tasks'

    
